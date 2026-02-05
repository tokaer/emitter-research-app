"""Parse Excel template into normalized input rows."""
from __future__ import annotations

import io
import logging
import re
from typing import Optional

from openpyxl import load_workbook
from unidecode import unidecode

from app.models import InputRowCreate

logger = logging.getLogger(__name__)

# Expected column headers (case-insensitive matching)
COLUMN_MAP = {
    "scope": "scope",
    "kategorie": "kategorie",
    "ggf. unterkategorie": "unterkategorie",
    "unterkategorie": "unterkategorie",
    "bezeichnung": "bezeichnung",
    "produktinformationen": "produktinformationen",
    "referenzeinheit": "referenzeinheit",
    "ggf. region": "region",
    "region": "region",
    "referenzjahr": "referenzjahr",
}


def _normalize_text(text: str) -> str:
    """Normalize text for search: lowercase, strip, collapse whitespace."""
    text = text.strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_for_search(text: str) -> str:
    """Normalize for search, including transliteration of umlauts."""
    return _normalize_text(unidecode(text))


def _normalize_region(region: Optional[str]) -> str:
    """Default region logic: empty/None -> GLO."""
    if not region or not region.strip():
        return "GLO"
    return region.strip().upper()


def parse_template(file_bytes: bytes) -> list[InputRowCreate]:
    """Parse an Excel template (.xlsx) into a list of InputRowCreate objects.

    Args:
        file_bytes: Raw bytes of the uploaded .xlsx file.

    Returns:
        List of InputRowCreate with mandatory fields validated.

    Raises:
        ValueError: If required columns are missing or no valid data rows found.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no active sheet")

    # Find column mapping from header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_indices: dict[str, int] = {}

    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        header_lower = str(cell_value).strip().lower()
        if header_lower in COLUMN_MAP:
            field_name = COLUMN_MAP[header_lower]
            col_indices[field_name] = idx

    # Validate required columns exist
    for required in ["bezeichnung", "referenzeinheit"]:
        if required not in col_indices:
            raise ValueError(
                f"Required column '{required}' not found in template headers. "
                f"Found headers: {[str(h).strip() for h in header_row if h]}"
            )

    # Parse data rows
    rows: list[InputRowCreate] = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        # Get cell values by mapped column index
        def get_val(field: str) -> Optional[str]:
            idx = col_indices.get(field)
            if idx is None or idx >= len(row_values):
                return None
            val = row_values[idx]
            if val is None:
                return None
            return str(val).strip()

        bezeichnung = get_val("bezeichnung")
        referenzeinheit = get_val("referenzeinheit")

        # Skip rows without mandatory fields
        if not bezeichnung or not referenzeinheit:
            continue

        row = InputRowCreate(
            scope=get_val("scope"),
            kategorie=get_val("kategorie"),
            unterkategorie=get_val("unterkategorie"),
            bezeichnung=bezeichnung,
            produktinformationen=get_val("produktinformationen"),
            referenzeinheit=referenzeinheit,
            region=get_val("region"),
            referenzjahr=get_val("referenzjahr"),
        )
        rows.append(row)

    wb.close()

    if not rows:
        raise ValueError("No valid data rows found in template (need Bezeichnung + Referenzeinheit)")

    logger.info(f"Parsed {len(rows)} rows from Excel template")
    return rows


def normalize_input_row(row: InputRowCreate) -> dict[str, str]:
    """Compute normalized fields for an input row.

    Returns dict with keys: bezeichnung_norm, produktinfo_norm, region_norm.
    """
    return {
        "bezeichnung_norm": _normalize_for_search(row.bezeichnung),
        "produktinfo_norm": _normalize_for_search(row.produktinformationen or ""),
        "region_norm": _normalize_region(row.region),
    }
